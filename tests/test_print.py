"""Tests for the PDF handout builders and their API endpoints."""
import re

import pytest
from fastapi.testclient import TestClient

from app.main import app
from app.print_pdf import (clip_view, overview_pdf, students_pdf,
                           teachers_pdf, term_label)
from app.scheduler import Dataset, Lesson, Room, Timeslot
from app.views import build_overview, build_student_view, build_teacher_view

MON, WED = "2026-07-27", "2026-07-29"


def make_data() -> Dataset:
    d = Dataset()
    d.students = {"s1": "葵", "s2": "蓮"}
    d.teachers = {"t1": "田中", "t2": "鈴木"}
    d.subjects = {"math": "数学", "eng": "英語"}
    d.rooms = {"r1": Room("r1", "第1教室", 2)}
    d.timeslots = {
        "mon-1": Timeslot("mon-1", MON, 1, "09:00-10:10"),
        "mon-2": Timeslot("mon-2", MON, 2, "10:20-11:30"),
        "wed-1": Timeslot("wed-1", WED, 1, "09:00-10:10"),
    }
    return d


LESSONS = [
    Lesson("s1", "math", "t1", "r1", "mon-1", id=1),
    Lesson("s2", "eng", "t2", "r1", "mon-1", id=2),
    Lesson("s2", "math", "t1", "r1", "wed-1", id=3),
]

STAMP = "2026-07-23 12:00"


def page_count(pdf: bytes) -> int:
    return max(int(n) for n in re.findall(rb"/Count (\d+)", pdf))


def test_overview_pdf_renders():
    d = make_data()
    out = overview_pdf(build_overview(d, LESSONS), STAMP)
    assert out.startswith(b"%PDF")
    assert page_count(out) == 1


def test_batch_pdfs_have_one_page_per_person():
    d = make_data()
    sv = [build_student_view(d, LESSONS, s) for s in ("s1", "s2")]
    out = students_pdf(sv, STAMP)
    assert out.startswith(b"%PDF")
    assert page_count(out) == 2
    tv = [build_teacher_view(d, LESSONS, t) for t in ("t1", "t2")]
    assert page_count(teachers_pdf(tv, STAMP)) == 2


def test_empty_batch_raises():
    with pytest.raises(ValueError):
        students_pdf([], STAMP)
    with pytest.raises(ValueError):
        teachers_pdf([], STAMP)


def test_term_label_spans_in_term_dates():
    d = make_data()
    assert term_label(build_overview(d, LESSONS)) == "7/27〜7/29"


def test_clip_view_restricts_date_range():
    d = make_data()
    v = build_overview(d, LESSONS)
    clipped = clip_view(v, None, MON)          # Monday only
    cells = {c["date"]: c for w in clipped["weeks"] for c in w}
    assert cells[MON]["in_term"] is True
    assert cells[WED]["in_term"] is False and cells[WED]["slots"] == []
    assert term_label(clipped) == "7/27〜7/27"
    # a range with no dates drops every week
    assert clip_view(v, "2030-01-01", "2030-12-31")["weeks"] == []


# ---------------------------------------------------------------- xlsx export

def test_overview_xlsx_structure():
    """The workbook mirrors the transposed table: header row, one
    n_sub-row block per day with a merged date cell, teacher|students
    column pairs, no subjects/rooms."""
    import io

    from openpyxl import load_workbook

    from app.export_xlsx import overview_xlsx
    d = make_data()
    out = overview_xlsx(build_overview(d, LESSONS), STAMP)
    assert out[:2] == b"PK"                      # zip container
    ws = load_workbook(io.BytesIO(out)).active
    assert ws.cell(row=1, column=1).value == "日付"
    assert ws.cell(row=1, column=2).value.startswith("①")
    # Monday block: two teachers in P1 (t1+t2), t1 again in P2
    assert ws.cell(row=2, column=1).value == "7/27(月)"
    p1 = {(ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value)
          for r in (2, 3)}
    assert p1 == {("田中", "葵"), ("鈴木", "蓮")}
    # n_sub = 2 -> Wednesday block starts at row 4; its date is merged
    assert ws.cell(row=4, column=1).value == "7/29(水)"
    assert str(ws.merged_cells.ranges).count("A") >= 1
    # subjects/rooms appear nowhere
    texts = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
    assert not any("数学" in t or "教室" in t for t in texts)


def test_batch_xlsx_one_sheet_per_person():
    """students/teachers workbooks carry one worksheet per person with
    the transposed day×period layout."""
    import io

    from openpyxl import load_workbook

    from app.export_xlsx import students_xlsx, teachers_xlsx
    d = make_data()
    sv = [build_student_view(d, LESSONS, s) for s in ("s1", "s2")]
    wb = load_workbook(io.BytesIO(students_xlsx(sv, STAMP)))
    assert wb.sheetnames == ["葵 (s1)", "蓮 (s2)"]
    ws = wb["蓮 (s2)"]
    assert ws.cell(row=1, column=2).value.startswith("①")
    assert ws.cell(row=2, column=1).value == "7/27(月)"
    assert ws.cell(row=2, column=2).value == "英語(鈴木)"   # s2 Mon P1
    assert ws.cell(row=3, column=2).value == "数学(田中)"   # s2 Wed P1

    tv = [build_teacher_view(d, LESSONS, t) for t in ("t1", "t2")]
    wb = load_workbook(io.BytesIO(teachers_xlsx(tv, STAMP)))
    assert wb.sheetnames == ["田中 (t1)", "鈴木 (t2)"]
    assert wb["田中 (t1)"].cell(row=2, column=2).value == "葵(数学)"

    import pytest as _pytest
    with _pytest.raises(ValueError):
        students_xlsx([], STAMP)


def test_batch_xlsx_endpoints(tmp_path):
    from fastapi.testclient import TestClient as TC
    app.state.db_path = tmp_path / "xlsx2.db"
    try:
        with TC(app) as c:
            seed(c)
            for url in ("/api/print/students.xlsx",
                        "/api/print/teachers.xlsx?ids=t1"):
                r = c.get(url)
                assert r.status_code == 200
                assert "spreadsheetml" in r.headers["content-type"]
                assert r.content[:2] == b"PK"
            assert c.get(
                "/api/print/students.xlsx?ids=ghost").status_code == 404
    finally:
        del app.state.db_path


def test_overview_xlsx_endpoint(tmp_path):
    from fastapi.testclient import TestClient as TC
    app.state.db_path = tmp_path / "xlsx.db"
    try:
        with TC(app) as c:
            seed(c)
            r = c.get("/api/print/overview.xlsx")
            assert r.status_code == 200
            assert "spreadsheetml" in r.headers["content-type"]
            assert r.content[:2] == b"PK"
            assert c.get(
                "/api/print/overview.xlsx?date_from=bad").status_code == 422
    finally:
        del app.state.db_path


# ------------------------------------------------------------------ endpoints

@pytest.fixture
def client(tmp_path):
    app.state.db_path = tmp_path / "print.db"
    with TestClient(app) as c:
        yield c
    del app.state.db_path


def seed(client):
    for path, body in [
        ("/api/students", {"id": "s1", "name": "Aoi"}),
        ("/api/teachers", {"id": "t1", "name": "Tanaka"}),
        ("/api/subjects", {"id": "math", "name": "Math"}),
        ("/api/rooms", {"id": "r1", "name": "Room 1"}),
        ("/api/timeslots", {"id": "mon-1", "date": MON, "period": 1}),
    ]:
        assert client.post(path, json=body).status_code == 200, path
    client.post("/api/lessons", json={
        "student_id": "s1", "subject_id": "math", "teacher_id": "t1",
        "room_id": "r1", "timeslot_id": "mon-1"})


@pytest.mark.parametrize("url", [
    "/api/print/overview.pdf",
    "/api/print/students.pdf",
    "/api/print/teachers.pdf",
    "/api/print/students.pdf?ids=s1",
    "/api/print/teachers.pdf?ids=t1&date_from=2026-07-01&date_to=2026-08-31",
])
def test_print_endpoints_return_pdf(client, url):
    seed(client)
    r = client.get(url)
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/pdf"
    assert r.content.startswith(b"%PDF")


def test_print_unknown_ids_404(client):
    seed(client)
    assert client.get("/api/print/students.pdf?ids=ghost").status_code == 404
    assert client.get("/api/print/teachers.pdf?ids=ghost").status_code == 404


def test_print_bad_date_422(client):
    seed(client)
    r = client.get("/api/print/overview.pdf?date_from=07/01")
    assert r.status_code == 422


def test_print_all_students_has_page_per_student(client):
    seed(client)
    client.post("/api/students", json={"id": "s2", "name": "Ren"})
    r = client.get("/api/print/students.pdf")
    assert page_count(r.content) == 2
