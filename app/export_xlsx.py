"""Excel (.xlsx) export of the transposed master table.

Same layout as the PDF overview handout (`print_pdf.overview_pdf`):
one block of rows per in-term day, one column pair per period
(student | teacher), each teacher holding a TWO sub-row lane across
the whole day — one student per row, the second left empty for a
single student — tinted with the teacher's UI color; a page break
after every week mirrors the PDF's week-per-page. Pure module:
consumes the views.py week-grid shape plus metadata and returns the
workbook bytes — no DB, no FastAPI.
"""
from __future__ import annotations

import datetime as dt
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

from .print_pdf import WEEKDAY_JA, _circled, teacher_fill_rgb, term_label

THIN = Side(style="thin", color="BBBBBB")
MEDIUM = Side(style="medium", color="000000")
GREY = PatternFill("solid", fgColor="EEEEEE")
HEAD_FILL = PatternFill("solid", fgColor="E3E3E3")
SUN = Font(color="BE1E1E")
SAT = Font(color="1E3CBE")
BOLD = Font(bold=True)


def _days_periods_labels(view: dict):
    days = [c for w in view["weeks"] for c in w if c["in_term"]]
    label_of: dict[int, str] = {}
    for d in days:
        for s in d["slots"]:
            if s["label"] and s["period"] not in label_of:
                label_of[s["period"]] = s["label"]
    return days, view["periods"], label_of


def _sheet_title(name: str, entity_id: str) -> str:
    """Excel-safe, unique worksheet title (31-char limit, no []:*?/\\)."""
    clean = "".join(ch for ch in name if ch not in "[]:*?/\\")
    suffix = f" ({entity_id})"
    return clean[:31 - len(suffix)] + suffix


def _page_setup(ws, view: dict, generated_at: str) -> None:
    ws.freeze_panes = "B2"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:1"
    ws.oddFooter.left.text = \
        f"期間 {term_label(view)} ・ 作成 {generated_at}"
    ws.oddFooter.right.text = "&P / &N ページ"


def _header_cell(cell) -> None:
    cell.font = BOLD
    cell.fill = HEAD_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _teacher_fill(teacher_id: str) -> PatternFill:
    r, g, b = teacher_fill_rgb(teacher_id)
    return PatternFill("solid", fgColor=f"{r:02X}{g:02X}{b:02X}")


def overview_xlsx(view: dict, generated_at: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "時間割 全体表"

    days, periods, label_of = _days_periods_labels(view)
    # week index of each in-term day, for the per-week print breaks
    week_of: list[int] = []
    for wi, w in enumerate(view["weeks"]):
        week_of += [wi] * sum(1 for c in w if c["in_term"])
    # sub-rows per teacher lane: 2 by design (one student per row,
    # capacity 2), stretched only if some teacher ever has more. Every
    # teacher working a day owns ONE lane across all its periods.
    rows_per = max([2] + [len(t["lessons"])
                          for d in days for s in d["slots"]
                          for t in s["entries"]])

    def day_teachers(day: dict) -> list[tuple[str, str]]:
        seen: dict[str, str] = {}
        for s in day["slots"]:
            for t in s["entries"]:
                seen.setdefault(t.get("teacher_id", t["teacher_name"]),
                                t["teacher_name"])
        return sorted(seen.items())

    # ---- header row: 日付 | ① 09:00-10:10 (merged over 2 cols) | …
    ws.cell(row=1, column=1, value="日付")
    _header_cell(ws.cell(row=1, column=1))
    for i, p in enumerate(periods):
        c0 = 2 + i * 2
        head = _circled(p)
        if label_of.get(p):
            head += f" {label_of[p]}"
        ws.merge_cells(start_row=1, start_column=c0,
                       end_row=1, end_column=c0 + 1)
        _header_cell(ws.cell(row=1, column=c0, value=head))

    # ---- one block of rows per day: one lane (rows_per rows) per
    # working teacher, blank in periods the teacher is off
    r = 2
    day_spans: list[tuple[int, int]] = []      # (first row, last row)
    for day in days:
        lanes = day_teachers(day)
        day_rows = max(1, len(lanes)) * rows_per
        day_spans.append((r, r + day_rows - 1))
        d = dt.date.fromisoformat(day["date"])
        wd = day["weekday"]
        ws.merge_cells(start_row=r, start_column=1,
                       end_row=r + day_rows - 1, end_column=1)
        date_cell = ws.cell(row=r, column=1,
                            value=f"{d.month}/{d.day}({WEEKDAY_JA[wd]})")
        date_cell.alignment = Alignment(horizontal="center",
                                        vertical="center")
        if wd == "Sun":
            date_cell.font = SUN
        elif wd == "Sat":
            date_cell.font = SAT
        slots = {s["period"]: s for s in day["slots"]}
        for i, p in enumerate(periods):
            c0 = 2 + i * 2
            slot = slots.get(p)
            here = ({t.get("teacher_id", t["teacher_name"]): t
                     for t in slot["entries"]} if slot else {})
            for row in range(day_rows):
                scell = ws.cell(row=r + row, column=c0)
                tcell = ws.cell(row=r + row, column=c0 + 1)
                if slot is None:
                    tcell.fill = GREY
                    scell.fill = GREY
                    continue
                k, j = divmod(row, rows_per)   # teacher lane, sub-row
                t = here.get(lanes[k][0]) if k < len(lanes) else None
                if t is not None:
                    fill = _teacher_fill(lanes[k][0])
                    tcell.fill = fill
                    scell.fill = fill
                    # the teacher's name on EVERY row of their lane,
                    # ONE student per row (second row empty when the
                    # teacher has a single student)
                    tcell.value = t["teacher_name"]
                    if j < len(t["lessons"]):
                        scell.value = t["lessons"][j]["student_name"]
        r += day_rows

    # ---- borders: thin inside a day×period cell, medium around it
    last_row = r - 1
    span_of = {row: span for span in day_spans
               for row in range(span[0], span[1] + 1)}
    for row in range(2, last_row + 1):
        block_top = span_of[row][0] == row
        block_bottom = span_of[row][1] == row
        ws.cell(row=row, column=1).border = Border(
            left=MEDIUM, right=MEDIUM,
            top=MEDIUM if block_top else THIN,
            bottom=MEDIUM if block_bottom else THIN)
        for i in range(len(periods)):
            c0 = 2 + i * 2
            ws.cell(row=row, column=c0).border = Border(
                left=MEDIUM, right=THIN,
                top=MEDIUM if block_top else THIN,
                bottom=MEDIUM if block_bottom else THIN)
            ws.cell(row=row, column=c0 + 1).border = Border(
                left=THIN, right=MEDIUM,
                top=MEDIUM if block_top else THIN,
                bottom=MEDIUM if block_bottom else THIN)
    for col in range(1, 2 + 2 * len(periods)):
        ws.cell(row=1, column=col).border = Border(
            left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

    # ---- one week per printed page (mirrors the PDF handout)
    for di in range(len(days) - 1):
        if week_of[di] != week_of[di + 1]:
            ws.row_breaks.append(Break(id=day_spans[di][1]))

    # ---- column widths, freeze panes, print setup
    ws.column_dimensions["A"].width = 10
    for i in range(len(periods)):
        ws.column_dimensions[get_column_letter(2 + i * 2)].width = 22
        ws.column_dimensions[get_column_letter(3 + i * 2)].width = 13
    _page_setup(ws, view, generated_at)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _calendar_sheet(ws, view: dict, generated_at: str,
                    cell_rows) -> None:
    """One month-style calendar: Mon-Sun columns, THREE rows per week —
    date / top content / bottom content. ``cell_rows(day_cell) ->
    (top, bottom)`` renders one day (both empty strings = free day)."""
    for i, wd in enumerate("月火水木金土日"):
        _header_cell(ws.cell(row=1, column=1 + i, value=wd))
    r = 2
    for week in view["weeks"]:
        for i, cell in enumerate(week):
            d = dt.date.fromisoformat(cell["date"])
            date_cell = ws.cell(row=r, column=1 + i,
                                value=f"{d.month}/{d.day}")
            wd = cell["weekday"]
            if wd == "Sun":
                date_cell.font = SUN
            elif wd == "Sat":
                date_cell.font = SAT
            if not cell["in_term"]:
                for rr in range(r, r + 3):
                    ws.cell(row=rr, column=1 + i).fill = GREY
                continue
            top, bottom = cell_rows(cell)
            if top:
                ws.cell(row=r + 1, column=1 + i, value=top)
            if bottom:
                ws.cell(row=r + 2, column=1 + i, value=bottom)
        r += 3
    # borders: medium around each day cell, thin between its 3 rows
    last_row = r - 1
    for row in range(2, last_row + 1):
        pos = (row - 2) % 3                      # 0 date, 1 top, 2 bottom
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = Border(
                left=MEDIUM, right=MEDIUM,
                top=MEDIUM if pos == 0 else THIN,
                bottom=MEDIUM if pos == 2 else THIN)
    for col in range(1, 8):
        ws.cell(row=1, column=col).border = Border(
            left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)
        ws.column_dimensions[get_column_letter(col)].width = 16
    _page_setup(ws, view, generated_at)
    ws.freeze_panes = "A2"                       # no sticky name column


def _student_cell_rows(cell: dict) -> tuple[str, str]:
    """Subjects on top, their periods (1限 style) below — comma-
    delimited, aligned by lesson order (same as the student PDF)."""
    entries = [(s["period"], e)
               for s in cell["slots"] for e in s["entries"]]
    entries.sort(key=lambda pe: pe[0])
    return ("、".join(e["subject_name"] for _p, e in entries),
            "、".join(f"{p}限" for p, _e in entries))


def students_xlsx(views: list[dict], generated_at: str) -> bytes:
    """One workbook, one CALENDAR worksheet per student: Mon-Sun
    columns, and per week a date row, a subjects row (comma-delimited)
    and a periods row (1限・2限 style) — the Excel twin of the student
    PDF handout."""
    if not views:
        raise ValueError("nothing to export")
    wb = Workbook()
    wb.remove(wb.active)
    for v in views:
        ws = wb.create_sheet(
            title=_sheet_title(v["student_name"], v["student_id"]))
        _calendar_sheet(ws, v, generated_at, _student_cell_rows)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _teacher_matrix_sheet(ws, view: dict, generated_at: str) -> None:
    """One week-matrix timetable: per week a block of rows — a date
    header (Mon-Sun columns) then one row per period (1限 style),
    each cell listing the teacher's students in that slot."""
    periods = view["periods"]
    label_of: dict[int, str] = {}
    for w in view["weeks"]:
        for c in w:
            for s in c["slots"]:
                if s["label"] and s["period"] not in label_of:
                    label_of[s["period"]] = s["label"]
    r = 1
    for week in view["weeks"]:
        head_row, first_period_row = r, r + 1
        _header_cell(ws.cell(row=r, column=1, value=""))
        for i, cell in enumerate(week):
            d = dt.date.fromisoformat(cell["date"])
            wd = cell["weekday"]
            hc = ws.cell(row=r, column=2 + i,
                         value=f"{d.month}/{d.day}({WEEKDAY_JA[wd]})")
            _header_cell(hc)
            if wd == "Sun":
                hc.font = SUN
            elif wd == "Sat":
                hc.font = SAT
        r += 1
        for p in periods:
            label = f"{p}限"
            if label_of.get(p):
                label += f" {label_of[p]}"
            lc = ws.cell(row=r, column=1, value=label)
            lc.alignment = Alignment(horizontal="center",
                                     vertical="center")
            for i, cell in enumerate(week):
                c = ws.cell(row=r, column=2 + i)
                has_slot = any(s["period"] == p for s in cell["slots"])
                if not cell["in_term"] or not has_slot:
                    c.fill = GREY
                    continue
                names = "、".join(
                    e["student_name"]
                    for s in cell["slots"] if s["period"] == p
                    for e in s["entries"])
                if names:
                    c.value = names
            r += 1
        # borders: medium frame around the week block, thin inside
        for row in range(head_row, r):
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = Border(
                    left=MEDIUM if col in (1, 2) else THIN,
                    right=MEDIUM if col == 8 else THIN,
                    top=MEDIUM if row in (head_row, first_period_row)
                    else THIN,
                    bottom=MEDIUM if row == r - 1 else THIN)
    ws.column_dimensions["A"].width = 14
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16
    _page_setup(ws, view, generated_at)
    ws.freeze_panes = "B1"                       # keep the period column


def teachers_xlsx(views: list[dict], generated_at: str) -> bytes:
    """One workbook, one week-matrix worksheet per teacher (period
    rows × day columns, students in the cells) — the Excel twin of the
    teacher PDF handout."""
    if not views:
        raise ValueError("nothing to export")
    wb = Workbook()
    wb.remove(wb.active)
    for v in views:
        ws = wb.create_sheet(
            title=_sheet_title(v["teacher_name"], v["teacher_id"]))
        _teacher_matrix_sheet(ws, v, generated_at)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
